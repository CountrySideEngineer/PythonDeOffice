import abc
import openpyxl
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.worksheet.header_footer import _HeaderFooterPart
import OfficeFile
from OfficeHeaderFooter import OfficeHeaderFooter
from HeaderFooterItem import HeaderFooterItem

class ExcelFile(OfficeFile.IOfficeFile):
	def __init__(self, path : str = ""):
		"""Constructor

		"""
		super().__init__(path = path)

	def Write(self, item : OfficeHeaderFooter) -> None:
		try:
			wb = openpyxl.load_workbook(self.path)
		except FileNotFoundError:
			print('Input file not found.')
		else:
			header_footer_item = self.ExportItem(src=item)

			for sheet_name in wb.sheetnames:
				ws = wb[sheet_name]
				try:
					left_part = self.GetLeftPartFromSheet(sheet=ws)
					left_item = header_footer_item[0]
					self.WriteIntoPart(part=left_part, item=left_item)

					center_part = self.GetCenterPartFromSheet(sheet=ws)
					center_item = header_footer_item[1]
					self.WriteIntoPart(part=center_part, item=center_item)

					right_part = self.GetRightPartFromSheet(sheet=ws)
					right_item = header_footer_item[2]
					self.WriteIntoPart(part=right_part, item=right_item)
				except IndexError as e:
					print('Index error detected while writing header/footer')
					print('Skip writing item and go to next sheet.')

	def WriteAll(self, items : list) -> None:
		try:
			wb = openpyxl.load_workbook(self.path)
		except FileNotFoundError:
			print('Input file not found.')
		else:
			for item in items:
				header_footer_item = self.ExportItem(src=item)
				try:
					ws = wb[item.name]
				except IndexError:
					print(f'{item.name} can not find.')
				else:
					self.WriteIntoSheet(ws=ws, header_footer_item=header_footer_item)

	def Read(self) -> list:
		wb = openpyxl.load_workbook(self.path)
		items = self.ReadFromBook(wb)

		return items

	def ReadFromBook(self, wb : Workbook) -> list:
		sheet_names = wb.sheetnames
		header_footers = []
		for sheet_name in sheet_names:
			ws = wb[sheet_name]
			hf = self.ReadFromSheet(ws)
			hf.name = sheet_name
			header_footers.append(hf)

		return header_footers

	def ReadFromSheet(self, ws : worksheet) -> OfficeHeaderFooter:
		left_part = self.GetLeftPartFromSheet(sheet = ws)
		left_item = HeaderFooterItem()
		left_item.item = self.ReadFromPart(part = left_part)

		center_part = self.GetCenterPartFromSheet(sheet = ws)
		center_item = HeaderFooterItem()
		center_item.item = self.ReadFromPart(part = center_part)

		right_part = self.GetRightPartFromSheet(sheet = ws)
		right_item = HeaderFooterItem()
		right_item.item = self.ReadFromPart(part = right_part)

		hf = OfficeHeaderFooter()
		self.AppendItem(dst = hf, item = left_item)
		self.AppendItem(dst = hf, item = center_item)
		self.AppendItem(dst = hf, item = right_item)

		return hf

	def ReadFromPart(self, part : _HeaderFooterPart) -> str:
		item = part.text
		return item

	def WriteIntoSheet(self, ws : worksheet, header_footer_item : list) -> None:
		try:
			left_part = self.GetLeftPartFromSheet(sheet=ws)
			left_item = header_footer_item[0]
			self.WriteIntoPart(part=left_part, item=left_item)

			center_part = self.GetCenterPartFromSheet(sheet=ws)
			center_item = header_footer_item[1]
			self.WriteIntoPart(part=center_part, item=center_item)

			right_part = self.GetRightPartFromSheet(sheet=ws)
			right_item = header_footer_item[2]
			self.WriteIntoPart(part=right_part, item=right_item)
		except IndexError as e:
			print('Index error detected while writing header/footer')
			print('Skip writing item and go to next sheet.')

	def WriteIntoPart(self, part : _HeaderFooterPart, item : str) -> None:
		part.text = item

	@abc.abstractmethod
	def AppendItem(self, dst : OfficeHeaderFooter, item : HeaderFooterItem) -> None:
		raise NotImplementedError()

	@abc.abstractmethod
	def ExportItem(self, src : OfficeHeaderFooter) -> list:
		raise NotImplementedError()

	@abc.abstractmethod
	def GetLeftPartFromSheet(self, sheet : worksheet) -> _HeaderFooterPart:
		raise NotImplementedError()

	@abc.abstractmethod
	def GetCenterPartFromSheet(self, sheet : worksheet) -> _HeaderFooterPart:
		raise NotImplementedError()

	@abc.abstractmethod
	def GetRightPartFromSheet(self, sheet : worksheet) -> _HeaderFooterPart:
		raise NotImplementedError()
