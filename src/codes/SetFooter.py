from ast import arg
from email import headerregistry
import sys
import os
import openpyxl
from openpyxl import worksheet
from openpyxl.worksheet.header_footer import _HeaderFooterPart
import JointText as jt

def SetFooterRight(path:str, headers:list) -> None:
	"""Set header of right.

	Setup right header of all sheets in a excel file.

	Args:
		path(str): Path to file to set header.
		headetrs(list): Collection of strings to set into header.
						One item is one line in header.
						When output, all items are joined by change line code.
	"""
	wb = openpyxl.load_workbook(path)

	sheets = wb.worksheets
	for sheet in sheets:
		SetPageFooterRight(sheet=sheet, headers=headers)

	wb.save(path)
	wb.close()

def SetPageFooterRight(sheet:worksheet, headers:list) -> None:
	"""Set header of right.

	Setup right header of all sheets in a excel file.

	Args:
		sheet(worksheet): Worksheet object to set header.
		headetrs(list): Collection of strings to set into header.
						One item is one line in header.
						When output, all items are joined by change line code.
	"""
	header_part = sheet.oddFooter.right
	SetFooter(footer_part=header_part, headers=headers)

def SetFooterLeft(path:str, headers:list) -> None:
	"""Set header of left.

	Setup left header of all sheets in a excel file.

	Args:
		path(str): Path to file to set header.
		headetrs(list): Collection of strings to set into header.
						One item is one line in header.
						When output, all items are joined by change line code.
	"""
	wb = openpyxl.load_workbook(path)

	sheets = wb.worksheets
	for sheet in sheets:
		SetPageFooterLeft(sheet=sheet, headers=headers)

	wb.save(path)
	wb.close()

def SetPageFooterLeft(sheet:worksheet, headers:list) -> None:
	"""Set header of left.

	Setup left header of all sheets in a excel file.

	Args:
		sheet(worksheet): Worksheet object to set header.
		headetrs(list): Collection of strings to set into header.
						One item is one line in header.
						When output, all items are joined by change line code.
	"""
	fooder_part = sheet.oddFooter.left
	SetFooter(footer_part=fooder_part, headers=headers)


def SetFooterCenter(path:str, headers:list) -> None:
	"""Set header of center.

	Setup center header of all sheets in a excel file.

	Args:
		path(str): Path to file to set header.
		headetrs(list): Collection of strings to set into header.
						One item is one line in header.
						When output, all items are joined by change line code.
	"""
	wb = openpyxl.load_workbook(path)

	sheets = wb.worksheets
	for sheet in sheets:
		SetPageFooterCenter(sheet=sheet, headers=headers)

	wb.save(path)
	wb.close()

def SetPageFooterCenter(sheet:worksheet, headers:list) -> None:
	"""Set header of center.

	Setup center header of all sheets in a excel file.

	Args:
		sheet(worksheet): Worksheet object to set header.
		headetrs(list): Collection of strings to set into header.
						One item is one line in header.
						When output, all items are joined by change line code.
	"""
	header_part = sheet.oddFooter.center
	SetFooter(footer_part=header_part, headers=headers)

def SetFooter(footer_part:_HeaderFooterPart, headers:list) -> None:
	"""Set header.

	Args:
		footer_part(_HeaderFooterPart): _HeaderFooterPart object to set the headers.
		headetrs(list): Collection of strings to set into header.
						One item is one line in header.
						When output, all items are joined by change line code.
	"""
	fooder_text = jt.JoinText(headers=headers)
	footer_part.text = fooder_text

if '__main__' == __name__:
	path = sys.argv[1]
	argc = len(sys.argv)
	headers = sys.argv[2:argc]
	SetFooterCenter(path=path, headers=headers)
